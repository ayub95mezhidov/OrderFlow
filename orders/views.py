from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from decimal import Decimal
from collections import defaultdict
from django.http import HttpResponseRedirect
from django.views.generic import ListView
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.utils import timezone
from collections import OrderedDict

from datetime import date

from .models import Order, Customer, Accessories
from .forms import CustomerForm, AddOrderForm, CustomerSearchForm, AccessoriesForm
from settings.models import PriceSettings, PriceAccessories
from finance.models import Wallet, Debt, Income, Profit

def welcome(request):
    return render(request, "orders/welcome.html")

# Список клиентов
class CustomerListView(LoginRequiredMixin, ListView):
    model = Customer
    template_name = 'orders/customers.html'
    context_object_name = 'customers'
    paginate_by = 20

    def get_queryset(self):
        search_query = self.request.GET.get('search_query', '').strip()
        queryset = Customer.objects.filter(user=self.request.user)

        if search_query:
            queryset = queryset.filter(
                Q(full_name__icontains=search_query) |
                Q(phone_number__icontains=search_query)
            )

        return queryset.order_by('full_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_form'] = CustomerSearchForm(
            initial={'search_query': self.request.GET.get('search_query', '')}
        )
        context['search_query'] = self.request.GET.get('search_query', '')
        return context

@login_required
def customers_with_debt(request):
    customers = Customer.objects.filter(user=request.user)

    list_customers = []
    for customer in customers:
        order = Order.objects.filter(user=request.user, customer=customer, payment_status='not paid')
        if len(order) != 0:
            list_customers.append(customer)

    context = {
        'customers': list_customers,
    }
    return render(request, 'orders/customers.html', context=context)


# Добавления клиента
@login_required(login_url='/users/login/')
def add_customer(request):
    if request.method == 'POST':
        form = CustomerForm(data=request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.user = request.user
            customer.save()
            return HttpResponseRedirect(reverse('customers'))
    else:
        form = CustomerForm()
    context = {'form': form,}
    return render(request, 'orders/add_customer.html', context=context)

# Обнавление данных клиента
@login_required
def customer_update(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id, user=request.user)

    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return redirect('customers')
    else:
        form = CustomerForm(instance=customer)

    context = {'form': form, 'customer': customer}
    return render(request, 'orders/customer_update.html', context)

# Заказы клиента
@login_required(login_url='/users/login/')
def orders(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id, user=request.user)

    # Фильтры из GET-параметров
    payment_filter = request.GET.get('payment_status', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Получаем заказы и сортируем по дате (новые сверху)
    orders_list = Order.objects.filter(user=request.user, customer=customer).order_by('-order_date')
    accessories_list = Accessories.objects.filter(user=request.user, customer=customer).order_by('-order_date')

    # Применяем фильтры
    if payment_filter:
        orders_list = orders_list.filter(payment_status=payment_filter)
        accessories_list = accessories_list.filter(payment_status=payment_filter)
    if status_filter:
        orders_list = orders_list.filter(status=status_filter)
    if date_from:
        orders_list = orders_list.filter(order_date__date__gte=date_from)
        accessories_list = accessories_list.filter(order_date__date__gte=date_from)
    if date_to:
        orders_list = orders_list.filter(order_date__date__lte=date_to)
        accessories_list = accessories_list.filter(order_date__date__lte=date_to)

    # Пагинация
    paginator = Paginator(orders_list, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Используем вместо orders_list только заказы текущей страницы
    orders_list = page_obj.object_list

    # Группируем заказы и комплектующие по дате
    grouped_orders = OrderedDict()
    total_amount = {}  # Общая сумма за каждую дату
    for order in orders_list:
        date_key = order.order_date.strftime('%d.%m.%y')  # Формат 03.05.25
        if date_key not in grouped_orders:
            grouped_orders[date_key] = {'orders': [], 'accessories': []}
            total_amount[date_key] = 0
        total_amount[date_key] += order.total_sum
        grouped_orders[date_key]['orders'].append(order)

    for accessory in accessories_list:
        date_key = accessory.order_date.strftime('%d.%m.%y')
        if date_key not in grouped_orders:
            grouped_orders[date_key] = {'orders': [], 'accessories': []}
            total_amount[date_key] = 0
        total_amount[date_key] += accessory.accessories_total
        grouped_orders[date_key]['accessories'].append(accessory)

    # 1. Получаем все заказы и комплектующие клиента
    orders = Order.objects.filter(customer=customer).select_related('user')
    accessories = Accessories.objects.filter(customer=customer).select_related('user')

    # 2. Получаем все PriceSettings для пользователей этих заказов
    user_ids = list(orders.values_list('user_id', flat=True).distinct()) + \
               list(accessories.values_list('user_id', flat=True).distinct())
    price_settings = PriceSettings.objects.filter(user_id__in=user_ids)

    # 3. Создаем словарь цен для быстрого доступа
    price_dict = defaultdict(dict)
    for ps in price_settings:
        price_dict[ps.user_id][(ps.fabric_size, ps.ceiling_type)] = ps.price_m2

    # 4. Вычисляем долг для каждого заказа и общий долг
    total_debt = Decimal('0')
    orders_with_debt = []
    accessories_with_debt = []

    # Обрабатываем основные заказы
    for order in orders:
        price = price_dict.get(order.user_id, {}).get(
            (order.fabric_size, order.ceiling_type), Decimal('0')
        )

        square = order.width * order.length
        order_total = square * price
        order_debt = order_total if order.payment_status == 'not paid' else Decimal('0')

        # Добавляем вычисленные поля к объекту заказа
        order.calculated_square = square
        order.calculated_price = price
        order.calculated_total = order_total
        order.calculated_debt = order_debt

        orders_with_debt.append(order)
        total_debt += order_debt

    # Обрабатываем комплектующие
    for accessory in accessories:
        accessory_debt = accessory.accessories_total if accessory.payment_status == 'not paid' else Decimal('0')
        accessory.calculated_debt = accessory_debt
        accessories_with_debt.append(accessory)
        total_debt += accessory_debt

    price_settings = PriceSettings.objects.all()

    context = {
        'title': 'OrderFlow - Заказы клиента',
        'grouped_orders': grouped_orders,
        'total_amount': total_amount,
        'customer': customer,
        'price_settings': price_settings,
        'debt': total_debt,
        'accessories_with_debt': accessories_with_debt,
        'page_obj': page_obj,
        'filters': {
            'payment_status': payment_filter,
            'status': status_filter,
            'date_from': date_from,
            'date_to': date_to,
        },
        'order_status_choices': Order.STATUS_CHOICES,
        'payment_status_choices': Order.PAYMENT_STATUS_CHOICES,
    }
    return render(request, 'orders/orders.html', context)

def _apply_date_filters(qs, date_from, date_to):
    if date_from:
        qs = qs.filter(order_date__date__gte=date_from)
    if date_to:
        qs = qs.filter(order_date__date__lte=date_to)
    return qs

def _build_grouped_orders(orders_qs, accessories_qs=None):
    grouped = OrderedDict()
    total_amount = {}
    for order in orders_qs:
        key = order.order_date.strftime('%d.%m.%y')
        if key not in grouped:
            grouped[key] = {'orders': [], 'accessories': []}
            total_amount[key] = 0
        total_amount[key] += order.total_sum
        grouped[key]['orders'].append(order)
    if accessories_qs:
        for acc in accessories_qs:
            key = acc.order_date.strftime('%d.%m.%y')
            if key not in grouped:
                grouped[key] = {'orders': [], 'accessories': []}
                total_amount[key] = 0
            total_amount[key] += acc.accessories_total
            grouped[key]['accessories'].append(acc)
    return grouped, total_amount

# Все заказы отфильтрованные по статусу НОВЫЕ
@login_required(login_url='/users/login/')
def new_orders_all_customers(request):
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    payment_filter = request.GET.get('payment_status', '')

    orders_list = Order.objects.filter(user=request.user, status='new').order_by('-order_date')
    orders_list = _apply_date_filters(orders_list, date_from, date_to)
    if payment_filter:
        orders_list = orders_list.filter(payment_status=payment_filter)

    paginator = Paginator(orders_list, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    grouped_orders, total_amount = _build_grouped_orders(page_obj.object_list)

    context = {
        'title': 'OrderFlow - Новые заказы',
        'grouped_orders': grouped_orders,
        'total_amount': total_amount,
        'page_obj': page_obj,
        'filters': {'date_from': date_from, 'date_to': date_to, 'payment_status': payment_filter, 'status': 'new'},
        'payment_status_choices': Order.PAYMENT_STATUS_CHOICES,
        'show_payment_filter': True,
    }
    return render(request, 'orders/new_orders_all.html', context)

# Все заказы отфильтрованные по статусу В РАБОТЕ
@login_required(login_url='/users/login/')
def in_progress_orders_all_customers(request):
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    payment_filter = request.GET.get('payment_status', '')

    orders_list = Order.objects.filter(user=request.user, status='in_progress').order_by('-order_date')
    orders_list = _apply_date_filters(orders_list, date_from, date_to)
    if payment_filter:
        orders_list = orders_list.filter(payment_status=payment_filter)

    paginator = Paginator(orders_list, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    grouped_orders, total_amount = _build_grouped_orders(page_obj.object_list)

    context = {
        'title': 'OrderFlow - Заказы в работе',
        'grouped_orders': grouped_orders,
        'total_amount': total_amount,
        'page_obj': page_obj,
        'filters': {'date_from': date_from, 'date_to': date_to, 'payment_status': payment_filter, 'status': 'in_progress'},
        'payment_status_choices': Order.PAYMENT_STATUS_CHOICES,
        'show_payment_filter': True,
    }
    return render(request, 'orders/new_orders_all.html', context)

# Все заказы отфильтрованные по статусу ЗАВЕРШЕННЫЕ
@login_required(login_url='/users/login/')
def completed_orders_all_customers(request):
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    payment_filter = request.GET.get('payment_status', '')

    orders_list = Order.objects.filter(user=request.user, status='completed').order_by('-order_date')
    orders_list = _apply_date_filters(orders_list, date_from, date_to)
    if payment_filter:
        orders_list = orders_list.filter(payment_status=payment_filter)

    paginator = Paginator(orders_list, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    grouped_orders, total_amount = _build_grouped_orders(page_obj.object_list)

    context = {
        'title': 'OrderFlow - Завершённые заказы',
        'grouped_orders': grouped_orders,
        'total_amount': total_amount,
        'page_obj': page_obj,
        'filters': {'date_from': date_from, 'date_to': date_to, 'payment_status': payment_filter, 'status': 'completed'},
        'payment_status_choices': Order.PAYMENT_STATUS_CHOICES,
        'show_payment_filter': True,
    }
    return render(request, 'orders/new_orders_all.html', context)

# Все заказы отфильтрованные по статусу НЕ ОПЛАЧЕННЫЕ
@login_required(login_url='/users/login/')
def not_paid_orders_all_customers(request):
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    orders_list = Order.objects.filter(user=request.user, payment_status='not paid').order_by('-order_date')
    accessories_list = Accessories.objects.filter(user=request.user, payment_status='not paid').order_by('-order_date')
    orders_list = _apply_date_filters(orders_list, date_from, date_to)
    accessories_list = _apply_date_filters(accessories_list, date_from, date_to)

    paginator = Paginator(orders_list, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    grouped_orders, total_amount = _build_grouped_orders(page_obj.object_list, accessories_list)

    context = {
        'title': 'OrderFlow - Не оплаченные заказы',
        'grouped_orders': grouped_orders,
        'total_amount': total_amount,
        'page_obj': page_obj,
        'filters': {'date_from': date_from, 'date_to': date_to, 'payment_status': 'not paid', 'status': ''},
        'show_payment_filter': False,
    }
    return render(request, 'orders/new_orders_all.html', context)

# Все заказы отфильтрованные по статусу ОПЛАЧЕННЫЕ
@login_required(login_url='/users/login/')
def paid_orders_all_customers(request):
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    orders_list = Order.objects.filter(user=request.user, payment_status='paid').order_by('-order_date')
    accessories_list = Accessories.objects.filter(user=request.user, payment_status='paid').order_by('-order_date')
    orders_list = _apply_date_filters(orders_list, date_from, date_to)
    accessories_list = _apply_date_filters(accessories_list, date_from, date_to)

    paginator = Paginator(orders_list, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    grouped_orders, total_amount = _build_grouped_orders(page_obj.object_list, accessories_list)

    context = {
        'title': 'OrderFlow - Оплаченные заказы',
        'grouped_orders': grouped_orders,
        'total_amount': total_amount,
        'page_obj': page_obj,
        'filters': {'date_from': date_from, 'date_to': date_to, 'payment_status': 'paid', 'status': ''},
        'show_payment_filter': False,
    }
    return render(request, 'orders/new_orders_all.html', context)

# Обновляет статус заказа
@require_POST
def update_order_status(request, order_id):
    """Обнавляет статус заказа"""
    try:
        order = Order.objects.get(id=order_id)
        new_status = request.POST.get('status')

        if new_status in dict(order.STATUS_CHOICES):
            if new_status == 'completed' and not order.completed_at:
                order.completed_at = timezone.now()
            order.status = new_status
            order.save()

            customer = order.customer
            debt = calculate_customer_debt(customer)

            response = HttpResponse(render_to_string('orders/orders.html', {
                'order': order,
                'customer': customer,
                'debt': debt
            }))

            # Добавляем заголовок для триггера обновления долга
            response['HX-Trigger'] = 'StatusChanged'
            return response
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Order not found'})

# Обновляет статус платежа заказа
@login_required
@require_POST
def update_payment_status(request, order_id):
    """Обновляет статус платежа закза потолков"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    wallet, _ = Wallet.objects.get_or_create(user=request.user)

    new_status = request.POST.get('payment_status')
    today = date.today()

    profit_amount = order.profit() # прибыль с заказа
    if new_status in dict(Order.PAYMENT_STATUS_CHOICES):
        order.payment_status = new_status
        order.save()
        # Добавляет к кошельку сумму заказа и отнимает от долга
        if new_status == 'paid':
            wallet.cash += order.total_sum
            wallet.save()

            # Добавляет к общей прибыли прибыль с заказа
            profit, created = Profit.objects.get_or_create(user=request.user, date=today, defaults={'total': profit_amount})
            if not created:
                profit.total += profit_amount
                profit.save()

            from finance.models import CategoryIncome
            obj, create = CategoryIncome.objects.get_or_create(user=request.user, title='Потолки', color='Синий')
            Income.objects.create(
                user=request.user,
                total_sum=order.total_sum,
                title=obj,
                date=timezone.now().date(),
            )
        else:
            # Если статус платежа меняется на "Не оплачено" то с "Кошелка" отнимаеся сумма заказа
            wallet.cash -= order.total_sum
            wallet.save()

            # Отнимает с общей прибыли прибыль с заказа
            try:
                profit = Profit.objects.get(user=request.user, date=today)
                profit.total -= profit_amount
                profit.save()
            except Profit.DoesNotExist:
                pass

            # Из истории дохода мы вычеслеям сумму заказа
            from finance.models import CategoryIncome
            obj, create = CategoryIncome.objects.get_or_create(user=request.user, title='Потолки', color='Синий')
            income = Income.objects.filter(user=request.user, title=obj).last()
            if income:
                income.total_sum -= order.total_sum
                income.save()

        # Подсчитывает долг клиента
        customer = order.customer
        debt = calculate_customer_debt(customer)


        response = HttpResponse(render_to_string('orders/orders.html', {
            'order': order,
            'customer': customer,
            'debt': debt
        }))

    # Добавляем заголовок для триггера обновления долга
        response['HX-Trigger'] = 'paymentStatusChanged'
        return response

# Обнавление статуса платежа комплектующих
@login_required
@require_POST
def update_payment_status_accessories(request, accessories_id):
    """Обновляет статус платежа закза комплектующих"""
    accessories = get_object_or_404(Accessories, id=accessories_id, user=request.user)
    wallet, _ = Wallet.objects.get_or_create(user=request.user)
    new_status = request.POST.get('payment_status_accessories')

    today = date.today()

    # Добавляет к общей прибыли прибыль с заказа
    profit_amount = accessories.profit()
    if accessories_id:
        if new_status in dict(Order.PAYMENT_STATUS_CHOICES):
            accessories.payment_status = new_status
            accessories.save()
            # Добавляет к кошельку сумму заказа
            if new_status == 'paid':
                wallet.cash += accessories.accessories_total
                wallet.save()

                # Добавляет к общей прибыли прибыль с заказа
                profit, created = Profit.objects.get_or_create(user=request.user, date=today, defaults={'total': profit_amount})
                if not created:
                    profit.total += profit_amount
                    profit.save()

                from finance.models import CategoryIncome
                obj, create = CategoryIncome.objects.get_or_create(user=request.user, title='Комплектующие', color='Красный')
                Income.objects.create(
                    user=request.user,
                    total_sum=accessories.accessories_total,
                    title=obj,
                    date=timezone.now().date(),
                )
            else:
                # Если статус платежа меняется на "Не оплачено" то с "Кошелка" отнимаеся сумма заказа
                wallet.cash -= accessories.accessories_total
                wallet.save()

                # Отнимает с общей прибыли прибыль с заказа
                try:
                    profit = Profit.objects.get(user=request.user, date=today)
                    profit.total -= profit_amount
                    profit.save()
                except Profit.DoesNotExist as ex:
                    print(ex)

                # Из истории дохода мы вычеслеям сумму заказа
                from finance.models import CategoryIncome
                obj, create = CategoryIncome.objects.get_or_create(user=request.user, title='Комплектующие', color='Красный')
                income = Income.objects.filter(user=request.user, title=obj).last()
                if income:
                    income.total_sum -= accessories.accessories_total
                    income.save()

            # Пересчитываем долг клиента
            customer = accessories.customer
            debt = calculate_customer_debt(customer)


            response = HttpResponse(render_to_string('orders/orders.html', {
                'accessories': accessories,
                'customer': customer,
                'debt': debt
            }))

        # Добавляем заголовок для триггера обновления долга
            response['HX-Trigger'] = 'paymentStatusChanged'
            return response


@login_required
def customer_debt(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id, user=request.user)
    debt = calculate_customer_debt(customer)
    return HttpResponse(f"{debt:.2f}")

def calculate_customer_debt(customer):
    total = Decimal('0')
    for order in Order.objects.filter(customer=customer, payment_status='not paid'):
        total += order.total_sum

    for accessories in Accessories.objects.filter(customer=customer, payment_status='not paid'):
        total += accessories.accessories_total
    return total

# Создает заказ внутри клиента
@login_required
def add_order(request, customer_id=None):
    customer = None
    if customer_id:
        customer = get_object_or_404(Customer, id=customer_id)

    if request.method == 'POST':
        form = AddOrderForm(request.user, request.POST, request.FILES)
        if form.is_valid():
            order = form.save(commit=False)
            if customer:
                order.user = request.user
                order.customer = customer
            order.save()
            # Присваиваем сумму заказа(total_sum) debt и сохраняем
            debt, _ = Debt.objects.get_or_create(user=request.user)
            debt.debt += order.total_sum
            debt.save()
            return redirect('orders', customer_id=order.customer.id)
    else:
        initial = {'customer': customer} if customer else {}
        form = AddOrderForm(request.user, initial=initial)
    context = {'form': form,
               'customer': customer,
               'calculated_fields': {
                   'square': form.instance.square if form.instance.pk else 0,
                   'total_sum': form.instance.total_sum if form.instance.pk else 0,
                   'balance': form.instance.balance if form.instance.pk else 0,
               }}
    return render(request, 'orders/add_order.html', context=context)

@login_required
def update_order(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    if request.method == 'POST':
        form = AddOrderForm(
            user=request.user,
            instance=order,
            data=request.POST,
            files=request.FILES,
        )
        if form.is_valid():
            form.save()
            return redirect('orders', customer_id=order.customer.id)
    else:
        form = AddOrderForm(user=request.user, instance=order)

    context = {'form': form, 'order': order}
    return render(request, 'orders/update_order.html', context)

# Добавить заказ с возможностью выбрать клиента
@login_required
def add_order_for_orders(request, customer_id=None):
    customer = None
    if customer_id:
        customer = get_object_or_404(Customer, id=customer_id, user=request.user)

    if request.method == 'POST':
        form = AddOrderForm(request.user, request.POST, request.FILES)
        if form.is_valid():
            order = form.save(commit=False)
            order.user = request.user

            # Если клиент был выбран в форме, используем его
            if form.cleaned_data.get('customer'):
                order.customer = form.cleaned_data['customer']
            # Иначе используем клиента из URL (если был передан)
            elif customer:
                order.customer = customer

            if order.customer:  # Проверяем, что клиент установлен
                order.save()
                # Присваиваем сумму заказа(total_sum) debt и сохраняем
                debt, _ = Debt.objects.get_or_create(user=request.user)
                debt.debt += order.total_sum
                debt.save()
                return redirect('new_orders_all')
            else:
                # Если клиент не выбран, добавляем ошибку
                form.add_error('customer', 'Выберите клиента')
    else:
        initial = {'customer': customer} if customer else {}
        form = AddOrderForm(request.user, initial=initial)

    context = {
        'form': form,
        'customer': customer,
        'calculated_fields': {
            'square': form.instance.square if form.instance.pk else 0,
            'total_sum': form.instance.total_sum if form.instance.pk else 0,
            'balance': form.instance.balance if form.instance.pk else 0,
        }
    }
    return render(request, 'orders/add_order.html', context=context)

@login_required
def add_order_accessories(request, customer_id=None):
    customer = None
    if customer_id:
        customer = get_object_or_404(Customer, id=customer_id, user=request.user)

    if request.method == 'POST':
        form = AccessoriesForm(request.user, request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            if customer:
                order.user = request.user
                order.customer = customer
            order.save()
            # Присваиваем сумму заказа(total_sum) debt и сохраняем
            debt, _ = Debt.objects.get_or_create(user=request.user)
            debt.debt += order.accessories_total
            debt.save()
            return redirect('orders', customer_id=order.customer.id)
    else:
        initial = {'customer': customer} if customer else {}
        form = AccessoriesForm(request.user, initial=initial)
    context = {'form': form,
               'customer': customer,
               }
    return render(request, 'orders/add_accessories.html', context=context)

@login_required
def orders_remove(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    order.delete()
    return HttpResponseRedirect(request.META['HTTP_REFERER'])

@login_required
def orders_acc_remove(request, order_id):
    order = get_object_or_404(Accessories, id=order_id, user=request.user)
    order.delete()
    return HttpResponseRedirect(request.META['HTTP_REFERER'])

@login_required
def customers_remove(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id, user=request.user)
    customer.delete()
    return HttpResponseRedirect(request.META['HTTP_REFERER'])


# ─── AI-сканирование чертежей ────────────────────────────────────────────────

@login_required
def scan_drawings(request, customer_id=None):
    from django.contrib import messages as django_messages
    from .ai_recognition import recognize_rooms

    customers = Customer.objects.filter(user=request.user).order_by('full_name')
    customer = None
    if customer_id:
        customer = get_object_or_404(Customer, id=customer_id, user=request.user)

    if request.method == 'POST':
        provider = request.POST.get('provider', 'claude')
        selected_customer_id = request.POST.get('customer_id') or customer_id
        photos = request.FILES.getlist('photos')

        if not selected_customer_id:
            django_messages.error(request, 'Выберите клиента.')
            return render(request, 'orders/scan_drawings.html', {
                'customers': customers, 'customer': customer
            })

        if not photos:
            django_messages.error(request, 'Загрузите хотя бы одно фото.')
            return render(request, 'orders/scan_drawings.html', {
                'customers': customers, 'customer': customer
            })

        all_rooms = []
        errors = []
        for photo in photos:
            try:
                rooms = recognize_rooms(photo, provider)
                all_rooms.extend(rooms)
            except ValueError as e:
                django_messages.error(request, str(e))
                return render(request, 'orders/scan_drawings.html', {
                    'customers': customers, 'customer': customer
                })
            except Exception as e:
                errors.append(f'{photo.name}: {str(e)}')

        if errors:
            django_messages.warning(request, 'Не удалось обработать некоторые фото: ' + '; '.join(errors))

        if not all_rooms:
            django_messages.warning(request, 'AI не смог распознать размеры на загруженных фото. Попробуйте другие фото.')
            return render(request, 'orders/scan_drawings.html', {
                'customers': customers, 'customer': customer
            })

        # Нумеруем комнаты если имена дублируются
        for i, room in enumerate(all_rooms):
            if not room.get('name'):
                room['name'] = f'Комната {i + 1}'

        request.session['scanned_rooms'] = all_rooms
        request.session['scan_customer_id'] = int(selected_customer_id)
        return redirect('scan_confirm')

    return render(request, 'orders/scan_drawings.html', {
        'customers': customers,
        'customer': customer,
    })


@login_required
def scan_confirm(request):
    from django.contrib import messages as django_messages

    rooms = request.session.get('scanned_rooms')
    customer_id = request.session.get('scan_customer_id')

    if not rooms or not customer_id:
        django_messages.warning(request, 'Сессия истекла. Загрузите фото заново.')
        return redirect('scan_drawings')

    customer = get_object_or_404(Customer, id=customer_id, user=request.user)

    FABRIC_SIZE_CHOICES = [
        ('short', '1.5м - 3.6м'),
        ('wide', '4.0м - 5.0м'),
        ('the widest', '5.8м'),
    ]
    CEILING_TYPE_CHOICES = [
        ('white mat', 'Белый мат'),
        ('colored mat', 'Цветной мат'),
        ('white gloss', 'Белый лак'),
        ('colored gloss', 'Цветной лак'),
        ('white sateen', 'Белый сатин'),
        ('colored sateen', 'Цветной сатин'),
        ('venice', 'Венеция'),
        ('sky', 'Небо'),
    ]

    if request.method == 'POST':
        debt, _ = Debt.objects.get_or_create(user=request.user)
        created_count = 0

        count = int(request.POST.get('room_count', 0))
        for i in range(count):
            if request.POST.get(f'delete_{i}'):
                continue
            try:
                width = Decimal(request.POST.get(f'width_{i}', '0'))
                length = Decimal(request.POST.get(f'length_{i}', '0'))
                fabric_size = request.POST.get(f'fabric_size_{i}', 'short')
                ceiling_type = request.POST.get(f'ceiling_type_{i}', 'white mat')
            except Exception:
                continue

            if width <= 0 or length <= 0:
                continue

            order = Order(
                customer=customer,
                user=request.user,
                width=width,
                length=length,
                fabric_size=fabric_size,
                ceiling_type=ceiling_type,
                status='new',
                payment_status='not paid',
            )
            order.save()

            debt.debt += order.total_sum
            debt.save()

            created_count += 1

        del request.session['scanned_rooms']
        del request.session['scan_customer_id']

        django_messages.success(request, f'Создано заказов: {created_count}.')
        return redirect('orders', customer_id=customer.id)

    # GET: подготовить данные для шаблона
    enriched_rooms = []
    for i, room in enumerate(rooms):
        w = room['width']
        l = room['length']
        enriched_rooms.append({
            'index': i,
            'name': room.get('name', f'Комната {i+1}'),
            'width': f"{w:.2f}",
            'length': f"{l:.2f}",
            'fabric_size': room.get('fabric_size', 'short'),
            'ceiling_type': room.get('ceiling_type', 'white mat'),
            'square': f"{w * l:.2f}",
        })

    return render(request, 'orders/scan_confirm.html', {
        'customer': customer,
        'rooms': enriched_rooms,
        'room_count': len(enriched_rooms),
        'fabric_size_choices': FABRIC_SIZE_CHOICES,
        'ceiling_type_choices': CEILING_TYPE_CHOICES,
    })


@login_required
def export_orders_excel(request, customer_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse as HR

    customer = get_object_or_404(Customer, id=customer_id, user=request.user)
    orders_qs = Order.objects.filter(user=request.user, customer=customer).order_by('-order_date')
    accessories_qs = Accessories.objects.filter(user=request.user, customer=customer).order_by('-order_date')

    wb = openpyxl.Workbook()

    title_font = Font(bold=True, size=13)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='2563EB')
    center = Alignment(horizontal='center')

    # ── Лист 1: Потолки ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Потолки'

    # Строка с именем клиента
    ws1.merge_cells('A1:K1')
    title_cell = ws1['A1']
    title_cell.value = f'Клиент: {customer.full_name}'
    title_cell.font = title_font
    title_cell.alignment = center

    headers = ['Дата', 'Ширина (м)', 'Длина (м)', 'Тип', 'м²', 'Периметр', 'Ткань', 'Цена (₽)', 'Сумма (₽)', 'Оплата', 'Статус']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row, order in enumerate(orders_qs, 3):
        ws1.cell(row=row, column=1, value=order.order_date.strftime('%d.%m.%Y'))
        ws1.cell(row=row, column=2, value=float(order.width))
        ws1.cell(row=row, column=3, value=float(order.length))
        ws1.cell(row=row, column=4, value=order.get_ceiling_type_display())
        ws1.cell(row=row, column=5, value=float(order.square))
        ws1.cell(row=row, column=6, value=float(order.perimeter))
        ws1.cell(row=row, column=7, value=order.get_fabric_size_display())
        ws1.cell(row=row, column=8, value=float(order.price()))
        ws1.cell(row=row, column=9, value=float(order.total_sum))
        ws1.cell(row=row, column=10, value='Оплачено' if order.payment_status == 'paid' else 'Не оплачено')
        ws1.cell(row=row, column=11, value=order.get_status_display())

    col_widths = [14, 12, 12, 16, 8, 10, 16, 12, 12, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Лист 2: Комплектующие ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('Комплектующие')

    # Строка с именем клиента
    ws2.merge_cells('A1:E1')
    title_cell2 = ws2['A1']
    title_cell2.value = f'Клиент: {customer.full_name}'
    title_cell2.font = title_font
    title_cell2.alignment = center

    headers2 = ['Дата', 'Комплектующие', 'Кол-во', 'Сумма (₽)', 'Оплата']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row, acc in enumerate(accessories_qs, 3):
        ws2.cell(row=row, column=1, value=acc.order_date.strftime('%d.%m.%Y'))
        ws2.cell(row=row, column=2, value=str(acc.accessories))
        ws2.cell(row=row, column=3, value=acc.quantity)
        ws2.cell(row=row, column=4, value=float(acc.accessories_total))
        ws2.cell(row=row, column=5, value='Оплачено' if acc.payment_status == 'paid' else 'Не оплачено')

    col_widths2 = [14, 30, 10, 14, 14]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    filename = f"orders_{customer.full_name.replace(' ', '_')}.xlsx"
    response = HR(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
